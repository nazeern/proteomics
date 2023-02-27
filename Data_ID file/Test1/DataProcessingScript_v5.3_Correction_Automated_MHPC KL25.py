"""
KineMed Automated Data Processing Script (Script 2 of 2 in Workflow)

This script will generate the following files/folders in the KineMed data 
processing pipeline:
    
    1) Data Filter
    2) Summary Table
    3) Analysis Summary
           
The script requires 3 specific files within this directory in order to be run:
    
    1) Compound Report file(s)
    2) Parameters Template file
    3) MIDA database files
    
In addition, when executed, the script will automatically copy the current 
directory to the KineMed network under the corresponding project leader, project
code, notebook code and year. The script will then automatically generate an email
to the project leader and research associate, and create a copy of all output files 
on the network to be used specifically for thre Atlas of Molecular Kinetics.

On 11/29/16 Kelvin Li added the hidden ability to use "HUMAN" in the minimum MIDA
setting in the Parameters Template File to set the minimum EMX at 0.03 instead of
the DEFAULT value of 0.04 (or any of the weird stringency values).

On 12/12/16 Kelvin Li added the hidden ability to use "HUMAN25" in the minimum MIDA
setting in the Parameters Template File to set the minimum EMX at 0.025 instead of
the DEFAULT value of 0.04 (or any of the weird stringency values).
    
"""

# Import all modules
from __future__ import division
import __main__
from xlrd import open_workbook,xldate_as_tuple
from openpyxl import load_workbook
import pandas as pd
import itertools
import math
import numpy as np
from collections import OrderedDict
import os
from datetime import date,datetime
import time
import shutil
#import win32com.client as win32
import urllib

# Sort Compound Report files
def sortFiles(inString):
    sortKey = inString[0][-5:]+inString[0][0:inString[0].find('-')]
    return sortKey

# Get sample names from Compound Report file names for sorting purposes, depending
# on operating system
def getSampleString(fileString):
    if os.name == 'nt': 
        fileName = fileString.split('\\')[-1]
    elif os.name == 'posix':
        fileName = fileString.split('/')[-1]
    return fileName[fileName.lower().find('_d')+2:fileName.lower().find('_ms')]+'-'+fileName[0:fileName.find('_')]

# Sort Compound Report files based on sample names within file names    
def sortFiles2(inString):
    sortKey = getSampleString(inString)
    return sortKey    

# Load parameters and samples from 'Parameters Template' file
def loadParameters(parameterFile):
    wb = open_workbook(parameterFile,on_demand=True)
    
    parametersData = wb.sheet_by_name('Parameters')
    
    rtDiff_input = parametersData.row_values(3,1)[0]
    totalAbund_input = parametersData.row_values(4,1)[0]
    mDiffCriteria_input = parametersData.row_values(5,1)
    silacMasses_input = 4 if parametersData.row_values(6,1)[0] == "All" else 2
    rmsError_input =  float(parametersData.row_values(7,1)[0])
    dbScore_input = float(parametersData.row_values(8,1)[0])
    basePkAbund_input = float(parametersData.row_values(9,1)[0])
    em0Upper_input = float(parametersData.row_values(10,1)[0])
    em0Lower_input = float(parametersData.row_values(11,1)[0])
    peptideSD_input =  float(parametersData.row_values(12,1)[0])
    isotopomerSD_input = float(parametersData.row_values(13,1)[0])
    useAllIsotopomers_input = True if parametersData.row_values(14,1)[0]=="Yes" else False
    combinePeptides_input = True if parametersData.row_values(15,1)[0]=="Yes" else False
    #minMIDA = float(parametersData.row_values(16,1)[0]) ## OLD STATIC PARAMETER
    minMIDA_input = parametersData.row_values(16,1)[0]
    upperSILAC = float(parametersData.row_values(17,1)[0])
    lowerSILAC = float(parametersData.row_values(18,1)[0])
    silacSD = float(parametersData.row_values(19,1)[0])
    fractions = True if parametersData.row_values(20,1)[0]=="Yes" else False
    saturation = float(parametersData.row_values(21,1)[0])
    offsetSlope = [i for i in parametersData.row_values(22,1) if i != '']
    offsetIntercept = [i for i in parametersData.row_values(23,1) if i != '']
    
    wb.unload_sheet('Parameters')

    sampleData = wb.sheet_by_name('Sample')
    
    #if  sampleData.row_values(3,1)[0] == '':
    #    date2 = ''
    #else:
    #    date_value =xldate_as_tuple(sampleData.row_values(3,1)[0],wb.datemode)
    #    date2 = date(*date_value[:3])
        
    instrument = sampleData.row_values(0,1)[0]
    projectLeader = sampleData.row_values(1,1)[0]
    processedBy = sampleData.row_values(2,1)[0]
    #submitDate = date2
    submitDate = sampleData.row_values(3,1)[0]
    projectCode = sampleData.row_values(4,1)[0]
    notebookCode = sampleData.row_values(5,1)[0]
    tissueFluid = sampleData.row_values(6,1)[0]
    prep = sampleData.row_values(7,1)[0]
    species = sampleData.row_values(8,1)[0].upper()[0]
    
    variableData = pd.read_excel(parameterFile,sheetname='Sample',skiprows=9,parse_cols = [1,2,3,4,5,6])
    wb.unload_sheet('Sample')
    variableData = variableData.ix[1:]

    variableData.columns = ['Sample Code','Notebook Code','Subject Code','Tissue & Fraction','Time & Treatment','Body Water']
    variableData = variableData.dropna(how='all').reset_index(drop=True)
    variableData['Subject Code'] = variableData['Subject Code'].map(str)
    variableData['Subject Code'] = [str(variableData['Subject Code'][i][0:4]) if str(variableData['Time & Treatment'][i]).find('Standard')!=-1 else ('0000'+str(variableData['Subject Code'][i][0:4]))[-4:] for i in xrange(len(variableData))]
    variableData['Sample Names'] = variableData['Time & Treatment'].map(str) + "-" + variableData['Subject Code']+species
    samples_input = variableData['Sample Names'].values
    bodywater_input = variableData['Body Water'].values * 1/100
    
    wb.release_resources()
    
    sampleANDbw = zip(samples_input,bodywater_input)
    
    sortedSampled = [x for (x,y) in sampleANDbw]
    sortedBW = [y for (x,y) in sampleANDbw]
    
    # Define functions to calculate 'Minimum MIDA for EMx' for individual
    # sample based on body water (**Revised by Bill Holmes**)
    def normalStringency(x):
        return round(-893.89*x**3 + 121.85*x**2 -7.058*x,3)
    
    def moderateStringency(x):
        return round(-749.34*x**3 + 111.27*x**2 -7.3935*x,3)
        
    def highStringency(x):
        return round(-690.75*x**3 + 107.31*x**2 -7.6624*x,3)
    
    if minMIDA_input.upper() == 'DEFAULT':
        minMIDA = [0.04]*len(sortedBW)
    elif minMIDA_input.upper() == 'NORMAL':
        minMIDA = map(normalStringency,sortedBW)
    elif minMIDA_input.upper() == 'MODERATE':
        minMIDA = map(moderateStringency,sortedBW)
    elif minMIDA_input.upper() == 'HIGH':
        minMIDA = map(highStringency,sortedBW)
    elif minMIDA_input.upper() == 'HUMAN':
        minMIDA = [0.03]*len(sortedBW)
    elif minMIDA_input.upper() == 'HUMAN25':
        minMIDA = [0.025]*len(sortedBW)
    
    return sortedSampled, sortedBW, rtDiff_input, totalAbund_input, mDiffCriteria_input, silacMasses_input, rmsError_input,dbScore_input,basePkAbund_input,em0Upper_input,em0Lower_input,peptideSD_input,isotopomerSD_input,useAllIsotopomers_input, combinePeptides_input,instrument, projectLeader, processedBy, submitDate, projectCode, notebookCode, tissueFluid, prep, minMIDA, upperSILAC, lowerSILAC, silacSD, fractions,minMIDA_input, saturation, offsetSlope, offsetIntercept

# Import individual Compound Report file and extract information
def importCompoundReport(fileName):
    wb = open_workbook(fileName,on_demand=True)
    sheet = wb.sheet_by_index(0)
    sheetName = wb.sheet_names()[0]
    
    fileDate = str(sheet.row_values(7,0)[3])[0:str(sheet.row_values(7,0)[3]).find(' ')]
    
    for row in xrange(sheet.nrows):
        if str(sheet.row_values(row,0)[0])=='MS Spectral Peaks':
            startRow = row
            break
            
    wb.release_resources()        

    compoundData = pd.read_excel(fileName,sheetname=sheetName,skiprows=startRow+1,skip_footer=1)
    compoundData = compoundData[compoundData['Acc#+Name']!='Acc#+Name']
    del compoundData['ppm']
    del compoundData['Formula']
    
    return compoundData, fileDate

# Extract accession number from compound report data
def getAccession(inputString):
    return str(inputString.split('+')[0])

# Extract protein name from compound report data    
def getProteinName(inputString):
    return str(inputString.split('+')[1])

# Extract mass from compound report data    
def getNeutralMass(inputValue):
    return round(inputValue,1)

# Extract species from compound report data    
def getSpecies(inputString):
    return str(inputString.split('_')[0][-1])

# Extract notebook code from compound report data
def getNotebookCode(inputString):
    return str(inputString.split('_')[1])

# Extract fraction from compound report data
def getFraction(inputString):
    return str(inputString.split('_')[2])

# Extract peptide sequence from compound report data    
def getAASequence(inputString):
    return str(inputString.split('+')[1].split('-')[0])

# Extract full peptide sequence from compound report data
def getAAStartSeq(inputString):
    return str(inputString.split('-')[0])

# Extract amino acid start number from compound report data
def getAAStart(inputString):
    return str(inputString.split('+')[0])
    
# Extract peptide modifications from compound report data
def getModifications(inputString):
    return str(inputString.split('-')[1])

# Extract sample name from compound report data
def getSampleName(inputString):
    return inputString.split('_D')[1].split('_MS')[0]+"-"+inputString.split('_')[0]
    
# Split data from Compound Report files
def splitData(allData):
    newList = []
    headerList = []
    for i in xrange(len(allData)):
        tempData = []
        tempData = allData[i].ix[5:,0:7]
        tempData.columns = ['Peptide m/z','Calc m/z','Ion Cluster','Charge','Abund','Ion','Saturated']
        tempData = tempData[tempData['Peptide m/z']!='m/z']
        tempData=tempData.reset_index(drop=True)
        newList.append(tempData)
        headerList.append(allData[i].ix[1,0:])

    return newList, headerList

# Parse Compound Report filenames and extract information
def parseFileName(fileString,fractionOption):
    speciesString = fileString[fileString.find('_')-1:fileString.find('_')]
    codeString = fileString[fileString.find('_')+1:fileString[fileString.find('_')+1:].find('_')+fileString.find('_')+1]
    if fractionOption == True:
        fractionString = fileString[fileString.find(codeString)+len(codeString)+1:fileString.lower().find('_d')]
    if fractionOption == False:
        fractionString = ''

    return speciesString, codeString, fractionString

# Parse headers in Compound Report files
def parseHeader(header,dictSample,takeFractions):
    headerList = list(header)
    accession = str(headerList[1][0:headerList[1].find('+')])
    protein = str(headerList[1][headerList[1].find('+')+1:])
    mass = headerList[2]
    roundedNeutralMass = round(mass,1)
    rt = headerList[3]
    rtDiff = headerList[4]
    fileName = str(headerList[5])
    species, code, fraction = parseFileName(fileName,takeFractions)
    sample = dictSample[fileName]
    
    fullSequence =str(headerList[6])
    aaSequence = fullSequence[fullSequence.find('+')+1:fullSequence.find('-')]
    aaStart = fullSequence[0:fullSequence.find('-')]
    modifications = fullSequence[fullSequence.find('-')+1:]
    scoreDB =  headerList[9]

    return [fileName, species, code, fraction, sample, accession, protein, aaSequence, aaStart, modifications, mass, roundedNeutralMass, rt, rtDiff, scoreDB]

# Calculate 'Mx' 
def calculateMX(dataFrame):
    massChargeList = [[dataFrame['Acc#+Name'][i],dataFrame['Mass'][i],dataFrame['z'][i]] for i in xrange(len(dataFrame))]
    
    splitList = [list(g) for k, g in itertools.groupby(massChargeList)]
    mList = []
    for i in xrange(len(splitList)):
        for j in xrange(len(splitList[i])):
            if splitList[i][0][0] < 2400 and j <= 4:
                mList.append('M'+str(j))
            elif splitList[i][0][0] >= 2400 and j <= 5:
                mList.append('M'+str(j))
            else:
                mList.append('')

    return mList

# Calculate '%M' 
def calculatePercentM(dataFrame):
    mSequence1 = ['M0', 'M1', 'M2', 'M3']
    mSequence2 = ['M0', 'M1', 'M2', 'M3','M4']

    positions = [x for x in [([i,4] if dataFrame['M'][i:i+len(mSequence1)].tolist()==mSequence1 and dataFrame['Mass'][i]<2400 else ([i,5] if dataFrame['M'][i:i+len(mSequence2)].tolist()==mSequence2 and dataFrame['Mass'][i]>=2400 else ['',''])) for i in xrange(len(dataFrame['M']))] if x!=['','']]
    percentM = [0]*len(dataFrame['M'])

    for i in xrange(len(positions)):
        if dataFrame['Mass'][positions[i][0]] < 2400:
            abundSum = sum(dataFrame['Abund'][positions[i][0]:positions[i][0]+4].tolist())
            percentM[positions[i][0]] = dataFrame['Abund'][positions[i][0]]/abundSum
            percentM[positions[i][0]+1] = dataFrame['Abund'][positions[i][0]+1]/abundSum
            percentM[positions[i][0]+2] = dataFrame['Abund'][positions[i][0]+2]/abundSum
            percentM[positions[i][0]+3] = dataFrame['Abund'][positions[i][0]+3]/abundSum          
        elif dataFrame['Mass'][positions[i][0]]  >= 2400:
            abundSum = sum(dataFrame['Abund'][positions[i][0]:positions[i][0]+5].tolist())
            percentM[positions[i][0]] = dataFrame['Abund'][positions[i][0]]/abundSum
            percentM[positions[i][0]+1] = dataFrame['Abund'][positions[i][0]+1]/abundSum
            percentM[positions[i][0]+2] = dataFrame['Abund'][positions[i][0]+2]/abundSum
            percentM[positions[i][0]+3] = dataFrame['Abund'][positions[i][0]+3]/abundSum 
            percentM[positions[i][0]+4] = dataFrame['Abund'][positions[i][0]+4]/abundSum

    return percentM

# Calculate 'M Ratio'
def calculateMRatio(dataFrame):
    mRatio = [round(math.fabs(dataFrame['%M'][i]-dataFrame['%M'][i-1]),7) if dataFrame['M'][i]!='M0' and dataFrame['%M'][i]!=0 else '' for i in xrange(len(dataFrame['M']))]
    
    return mRatio

# Calculate 'Warnings' to remove 'BAD' peptides
def identifyWarnings(dataFrame,mDiff,abundFilter):
    
    maxDiff1 = mDiff[0]
    maxDiff2 = mDiff[1]
    maxDiff3 = mDiff[2]
    maxDiff4 = mDiff[3]
    
    maxSum = abundFilter
    warnings = ['OK']*len(dataFrame['M'])
    mSequence1 = ['M0', 'M1', 'M2', 'M3']
    mSequence2 = ['M0', 'M1', 'M2', 'M3','M4']
    
    positions = [x for x in [([i,4] if dataFrame['M'][i:i+len(mSequence1)].tolist()==mSequence1 and dataFrame['Mass'][i]<2400 else ([i,5] if dataFrame['M'][i:i+len(mSequence2)].tolist()==mSequence2 and dataFrame['Mass'][i]>=2400 else ['',''])) for i in xrange(len(dataFrame['M']))] if x!=['','']]
    
    for i in xrange(len(positions)):
        roundedMass = round(dataFrame['Mass'][positions[i][0]],1)
        maxRatios = max(dataFrame['M Ratio'][positions[i][0]+1:positions[i][0]+positions[i][1]].tolist())
        abundSum = sum(dataFrame['Abund'][positions[i][0]:positions[i][0]+positions[i][1]].tolist())

        if dataFrame['Sample'][positions[i][0]].upper().find('SILAM')!=-1 or dataFrame['Sample'][positions[i][0]].upper().find('SILAC')!=-1:
            warnings[positions[i][0]:positions[i][0]+positions[i][1]] = ['OK']*5 if positions[i][1]==5 else ['OK']*4
        elif roundedMass < 900 and maxRatios > maxDiff1:
            warnings[positions[i][0]:positions[i][0]+positions[i][1]] = ['BAD']*5 if positions[i][1]==5 else ['BAD']*4
        elif roundedMass > 899.99 and roundedMass < 1100 and maxRatios > maxDiff2:
            warnings[positions[i][0]:positions[i][0]+positions[i][1]] = ['BAD']*5 if positions[i][1]==5 else ['BAD']*4
        elif roundedMass > 1099.99 and roundedMass < 1800 and maxRatios > maxDiff3:
            warnings[positions[i][0]:positions[i][0]+positions[i][1]] = ['BAD']*5 if positions[i][1]==5 else ['BAD']*4
        elif roundedMass > 1799.99 and maxRatios > maxDiff4:
            warnings[positions[i][0]:positions[i][0]+positions[i][1]] = ['BAD']*5 if positions[i][1]==5 else ['BAD']*4
        elif roundedMass >=2400 and positions[i][1]==4:
            warnings[positions[i][0]:positions[i][0]+positions[i][1]] = ['NO M4']*4
        elif abundSum < maxSum:
            warnings[positions[i][0]:positions[i][0]+positions[i][1]] = ['LOW']*5 if positions[i][1]==5 else ['LOW']*4
        else:
            warnings[positions[i][0]:positions[i][0]+positions[i][1]] = ['OK']*5 if positions[i][1]==5 else ['OK']*4

    return warnings

# Calculate new columns to be inserted in Converted Compound Report 
def calculateNewColumns(massList,mList,abundList,sampleList,diffList,taFilter):
    percentMData = calculatePercentM(massList,mList,abundList)

    mRatioData = calculateMRatio(mList,percentMData)

    warningData = identifyWarnings(massList,mList,abundList,mRatioData,sampleList,diffList,taFilter)

    return percentMData, mRatioData, warningData

# Convert Compound Reports to data frame
def convertReport(peptideData,sDict,mdList,taList,useFractions):
    peptideData.insert(1,'Species',map(getSpecies,peptideData['FileName'].tolist()))
    peptideData.insert(2,'Code',map(getNotebookCode,peptideData['FileName'].tolist()))
    if useFractions == True:
        peptideData.insert(3,'Fraction',map(getFraction,peptideData['FileName'].tolist()))
    else:
        peptideData.insert(3,'Fraction',[np.nan]* len(peptideData['FileName'].tolist()))
    peptideData.insert(4,'Sample',map(getSampleName,peptideData['FileName'].tolist()))
    peptideData.insert(5,'Accession#',map(getAccession,peptideData['Acc#+Name'].tolist()))
    peptideData.insert(6,'Protein',map(getProteinName,peptideData['Acc#+Name'].tolist()))
    peptideData.insert(6,'AAstart+seq',map(getAAStartSeq,peptideData['Aastart+Sequ-Mods'].tolist()))
    peptideData.insert(7,'AASequence',map(getAASequence,peptideData['Aastart+Sequ-Mods'].tolist()))
    peptideData.insert(8,'AAStart',map(getAAStart,peptideData['Aastart+Sequ-Mods'].tolist()))
    peptideData.insert(9,'Modifications',map(getModifications,peptideData['Aastart+Sequ-Mods'].tolist()))
    
    peptideData.insert(11,'Rounded Neutral Mass',map(getNeutralMass,peptideData['Mass'].tolist()))

    # Calculate 'Mx' columns
    mxList = calculateMX(peptideData)
    peptideData.insert(15,'M',mxList)
    
    # Calculate '%M' column
    percentMData = calculatePercentM(peptideData)
    peptideData.insert(16,'%M',percentMData)
    
    # Calculate 'M Ratio ' column
    mRatioData = calculateMRatio(peptideData)
    peptideData.insert(17,'M Ratio',mRatioData)
    
    # Calculate 'Warnings' column
    warningList = identifyWarnings(peptideData,mdList,taList)
    peptideData.insert(18,'Warning',warningList)
    
    # Filter data frame to exclude any peptides with '%M' < 0
    peptideData = peptideData[peptideData['%M']>0]
    peptideData.fillna(0, inplace=True)
    
    del peptideData['AASequence']

    # Filter data frame to only include columns of interest and rename columns
    peptideData = peptideData[['FileName','Species','Code','Fraction','Sample','Accession#','Protein','AAstart+seq','Modifications','Mass','Rounded Neutral Mass','RT','RtDiff','Score(DB)','M','m/z','IsotopeCluster','z','Abund','%M','Saturated','Warning']]
    peptideData.columns = ['File Name','Species','Code','Fraction','Sample','Accession#','Protein','AAstart+seq','Modifications','Mass','Rounded Neutral Mass','Rt','RtDiff','Score (DB)','M','Peptide m/z','Ion Cluster','Charge','Abund','%M','Saturated','Warning']
    return peptideData

# Import all necessary files in directory
def getFiles(source):
    
    parametersFile = []
    midaFiles = []
    crFiles = []
    
    # Parse local directory and find files
    for dirpath, dirs, files in os.walk(source):
        for filename in files:
            fileName, fileExtension = os.path.splitext(filename)
            if fileName.find('CompoundReport')!=-1:
                crFiles.append(os.path.join(dirpath,filename))
            elif fileName.find('Parameters')!=-1:
                parametersFile.append(os.path.join(dirpath,filename))
            elif fileName.find('_DB')!=-1 and fileExtension=='.csv':
                midaFiles.append(os.path.join(dirpath,filename))
                    
    # Provide processor with warnings if necessary files are not present
    if len(crFiles)==0:
        print '*** NO COMPOUND REPORTS IN DIRECTORY! ***'
        raw_input('HIT ENTER TO EXIT...')
    elif len(parametersFile)==0:
        print '*** NO PARAMETERS TEMPLATE IN DIRECTORY! ***'
        raw_input('HIT ENTER TO EXIT...')
    elif len(midaFiles)==0:
        print '*** NO MIDA DATABASE IN DIRECTORY! ***'
        raw_input('HIT ENTER TO EXIT...')
    
    return parametersFile[0], midaFiles,crFiles

# Convert all Comound Report files and append to one table
def convertCompoundReport(crReports,dict1,var1, var2,useFractions):
    compiledReports = []
    
    for i in xrange(len(crReports)):
        report, reportDate = importCompoundReport(crReports[i])
        compiledReports.append(report)
        
    allReports = pd.concat(compiledReports)
    allReports = allReports.reset_index(drop=True)
    outputTab = convertReport(allReports,dict1,var1, var2,useFractions)
    
    sortedTab = outputTab.sort_index(by=['Protein','AAstart+seq','Charge','M','Sample'],ascending = [True,True,True,True,True])
    sortedTab['Saturated'] = sortedTab['Saturated'].replace('S',np.nan)
    sortedTab = sortedTab.reset_index(drop=True)
    sortedTab[['Abund','Saturated']] = sortedTab[['Abund','Saturated']].astype(float)
    
    sortedTab = sortedTab[sortedTab['Warning']=='OK']
    sortedTab = sortedTab.reset_index(drop=True)
    sortedTab.to_csv('SORTED.csv')

    sortedTab2 = pd.read_csv('SORTED.csv')
    pivotData = sortedTab2.pivot_table(['Mass','Rt','RtDiff','Score (DB)','Abund','%M','Saturated'],rows = ['Warning','Species','Protein','Accession#','AAstart+seq','Modifications','Charge','Fraction','Code','M'], cols = 'Sample')
    newMasses = [np.nanmin(pivotData.Mass.iloc[i].values) for i in xrange(len(pivotData.Mass))]
    
    del pivotData['Mass']
    os.remove('SORTED.csv')
    pivotData.insert(0,'Mass ',newMasses)
    
    return outputTab, pivotData, sortedTab2, reportDate

# Merge multiple MIDA databases into one
def condenseFiles(mida_file):
    tempMIDA = pd.read_csv(mida_file[0])
    for i in xrange(len(mida_file[1:])):
        tempMIDA = tempMIDA.append(pd.read_csv(mida_file[1:][i]))
    return tempMIDA

# Generate Summary File        
def generateSummaryFile(sampleList,midaDB,sortedTab,bodyWaterInput,silacMasses,satFilter,offsetSlope,offsetIntercept):
    
    # Get MIDA database(s) and filter
    midaTable = condenseFiles(midaDB)
    midaTable = midaTable.sort(columns=['Formula',' Cpd','Notes',' Mass',' RT'])
    midaTable=midaTable.drop_duplicates(cols=['Formula',' Cpd','Notes',' Mass'],take_last=True).reset_index()
    
    midaTable = midaTable.drop(['sequence','modifications','score','composition'],axis=1)
    midaTable[' Cpd'] = midaTable[' Cpd'].astype(object)
    midaTable = midaTable[midaTable[' Cpd'].map(len)>6]
    
    midaNames = [midaTable[' Cpd'].iloc[i][7:] for i in xrange(len(midaTable))]
    midaTable.insert(0,'Protein',midaNames)
    midaTable = midaTable.drop([' Cpd'],axis=1).reset_index().rename(columns = {'Notes': 'Sequence',' Mass':'Mass'})

    midaTable['Sequence'] = [midaTable['Sequence'][i].rstrip() for i in xrange(len(midaTable))]
    # COPY RESULTS FROM COMPOUND REPORT
    cpdReport = sortedTab.copy()
    
    newSequence = ['-'.join([cpdReport['AAstart+seq'].iloc[i].rstrip(),cpdReport['Modifications'].iloc[i].rstrip()]) for i in xrange(len(cpdReport))]
    cpdReport.insert(4,'Sequence',newSequence)
    cpdReport = cpdReport.drop(['Mass'],axis=1) ### REV 02.17.2015 ###
    
    # Merge Compound Report data frame and MIDA data frame
    
    mergedData = pd.merge(cpdReport,midaTable, on=['Protein','Sequence'], how ='outer')
    mergedData = mergedData[pd.notnull(mergedData['Code'])]
    if len(list(set(mergedData['Code']))) > 1:		
        mergedData['Code'] = [mergedData['Code'][0]]*len(mergedData)

    pivotData2 = mergedData.pivot_table(['Mass','Rt','RtDiff','Score (DB)','Abund','%M','Saturated'],rows = ['Species','Code','Protein','Accession#','AAstart+seq','Modifications','n','Charge','Fraction',' RT','Formula','M0', 'M1', 'M2', 'M3', 'M4', 'EM0 cubic coeff 3', 'EM0 cubic coeff 2', 'EM0 cubic coeff 1', 'EM1 cubic coeff 3', 'EM1 cubic coeff 2', 'EM1 cubic coeff 1', 'EM2 cubic coeff 3', 'EM2 cubic coeff 2', 'EM2 cubic coeff 1', 'EM3 cubic coeff 3', 'EM3 cubic coeff 2', 'EM3 cubic coeff 1', 'EM4 cubic coeff 3', 'EM4 cubic coeff 2', 'EM4 cubic coeff 1', 'All Swissprot IDs', 'multiple IDs (1=yes)'], cols = ['Sample','M'])
    massCheck = pivotData2.Mass.min(axis=1,skipna=True).values > 2400

    # Calculate Baseline Mx
    baselineM0 = np.asarray((pivotData2.reset_index().iloc[:,11].values/np.where(massCheck,pivotData2.reset_index().iloc[:,11:16].sum(axis=1),pivotData2.reset_index().iloc[:,11:15].sum(axis=1))))
    baselineM1 = np.asarray((pivotData2.reset_index().iloc[:,12].values/np.where(massCheck,pivotData2.reset_index().iloc[:,11:16].sum(axis=1),pivotData2.reset_index().iloc[:,11:15].sum(axis=1))))
    baselineM2 = np.asarray((pivotData2.reset_index().iloc[:,13].values/np.where(massCheck,pivotData2.reset_index().iloc[:,11:16].sum(axis=1),pivotData2.reset_index().iloc[:,11:15].sum(axis=1))))
    baselineM3 = np.asarray((pivotData2.reset_index().iloc[:,14].values/np.where(massCheck,pivotData2.reset_index().iloc[:,11:16].sum(axis=1),pivotData2.reset_index().iloc[:,11:15].sum(axis=1))))
    baselineM4 = np.asarray((pivotData2.reset_index().iloc[:,15].values/np.where(massCheck,pivotData2.reset_index().iloc[:,11:16].sum(axis=1),pivotData2.reset_index().iloc[:,11:15].sum(axis=1))))
    
    def multiplyBW(inData):
        return map(np.sum,inData* np.asarray(bodyWaterInput)[np.newaxis].T**np.asarray([3,2,1]))
        
    # Calculate MIDA Mx
    midaEM0 =  map(multiplyBW,pivotData2.reset_index().iloc[:,16:19].values)
    midaEM1 =  map(multiplyBW,pivotData2.reset_index().iloc[:,19:22].values)
    midaEM2 =  map(multiplyBW,pivotData2.reset_index().iloc[:,22:25].values)
    midaEM3 =  map(multiplyBW,pivotData2.reset_index().iloc[:,25:28].values)
    midaEM4 =  map(multiplyBW,pivotData2.reset_index().iloc[:,28:31].values)
    
    summaryTable = pivotData2.reset_index().ix[0:,0:33]
    
    summaryTable.columns = [str(i[0]) for i in summaryTable.columns]
    
    summaryTable = summaryTable.rename(columns = {' RT':'Target RT'})
    
    def redefineSamples(samples):
        seen=set()
        seen_add = seen.add
        return [x for x in samples if not (x in seen or seen_add(x))]
    sampleWaterData = OrderedDict((x,y) for (x,y) in zip(sampleList,bodyWaterInput))
    
    sampleList = redefineSamples([i for i in sampleList for j in pivotData2.columns if j[1].find(i)!=-1])
    bwList = [sampleWaterData[i] for i in sampleList]
    
    # Calculate Base Peak Abundance
    basePkAbund = map(list,zip(*[pivotData2['Abund',sampleList[i]].max(axis=1,skipna=True).values for i in xrange(len(sampleList))]))
    
    # Calculate Sample Score
    summaryTable.insert(0,'Sample Score',map(np.count_nonzero,map(np.nan_to_num,basePkAbund)))
    
    def calcAbundScore(a):
        return 1 if a < 60000 else (2 if a < 120000 else (3 if a < 200000 else 4))
    
    # Calculate Abundance Score
    summaryTable.insert(1,'Abund Score',map(calcAbundScore,[c/t for c,t in zip(map(np.nansum,basePkAbund), summaryTable['Sample Score'].values)]))
    
    # Calculate Sample * Abundance Score
    summaryTable.insert(2, 'Sample*Abund Score',summaryTable['Sample Score'] * summaryTable['Abund Score'])
    
    # Insert additional columns in data frame for Corrected Mx and Mass
    summaryTable.insert(9,'Mass',map(np.nanmin,pivotData2.Mass.values))
    summaryTable.insert(20,'M0_c',baselineM0)
    summaryTable.insert(21,'M1_c',baselineM1)
    summaryTable.insert(22,'M2_c',baselineM2)
    summaryTable.insert(23,'M3_c',baselineM3)
    summaryTable.insert(24,'M4_c',baselineM4)
    
    summaryTable2 = summaryTable.ix[0:,25:]
    summaryTable = summaryTable.ix[0:,0:25]
    
    def calcSumAbundALL(inData):
        return np.nansum(inData)

    def calcSumAbund2(inData):
        return np.nansum(inData[0:2])
    
    # Calculate Summed Abundance
    sumAbund = [map(calcSumAbund2,pivotData2['Abund',sampleList[i]].values) if sampleList[i].find('SILAC')!=-1 and silacMasses==2 else map(calcSumAbundALL,pivotData2['Abund',sampleList[i]].values) for i in xrange(len(sampleList))]
    
    def calcSat(a):
        return 1 if a >satFilter else np.nan
    
    # Calculate Saturation
    sat = [map(calcSat,pivotData2['Abund',sampleList[i]].max(axis=1,skipna=True).values) for i in xrange(len(sampleList))]
    
    # Calculate EMx
    EM0 = map(list,zip(*[[a - b for a, b in zip(list(zip(*pivotData2['%M',sampleList[i]].values)[0]), baselineM0)] for i in xrange(len(sampleList))]))
    EM1 = map(list,zip(*[[a - b for a, b in zip(list(zip(*pivotData2['%M',sampleList[i]].values)[1]), baselineM1)] for i in xrange(len(sampleList))]))
    EM2 = map(list,zip(*[[a - b for a, b in zip(list(zip(*pivotData2['%M',sampleList[i]].values)[2]), baselineM2)] for i in xrange(len(sampleList))]))
    EM3 = map(list,zip(*[[a - b for a, b in zip(list(zip(*pivotData2['%M',sampleList[i]].values)[3]), baselineM3)] for i in xrange(len(sampleList))]))
    
    if 'M4' in list(set(cpdReport['M'].tolist())):
        EM4 = map(list,zip(*[[a - b for a, b in zip(list(zip(*pivotData2['%M',sampleList[i]].values)[4]), baselineM4)] if ('%M',sampleList[i],'M4') in pivotData2.columns else [0 for a, b in zip(list(zip(*pivotData2['%M',sampleList[0]].values)[0]), baselineM4)] for i in xrange(len(sampleList))]))
    else:
        EM4 = map(list,zip(*[[0 for a, b in zip(list(zip(*pivotData2['%M',sampleList[0]].values)[0]), baselineM0)] for i in xrange(len(sampleList))]))
    
    # Calculate Theoretical Mx
    theoEM0 = [[c/t if t!=0 else np.nan for c,t in zip(EM0[i],midaEM0[i])] for i in xrange(len(EM0))]
    theoEM1 = [[c/t if t!=0 else np.nan for c,t in zip(EM1[i],midaEM1[i])] for i in xrange(len(EM1))]
    theoEM2 = [[c/t if t!=0 else np.nan for c,t in zip(EM2[i],midaEM2[i])] for i in xrange(len(EM2))]
    theoEM3 = [[c/t if t!=0 else np.nan for c,t in zip(EM3[i],midaEM3[i])] for i in xrange(len(EM3))]
    if 'M4' in list(set(cpdReport['M'].tolist())):
        theoEM4 = [[c/t if t!=0 else np.nan for c,t in zip(EM4[i],midaEM4[i])] for i in xrange(len(EM4))]
    #else:
    #    theoEM4 = ['']*len(theoEM0)

    # Calculate Observed Mx
    M0obs = [map(lambda x:x+baselineM0[i], EM0[i]) for i in xrange(len(EM0))]
    M1obs = [map(lambda x:x+baselineM1[i], EM1[i]) for i in xrange(len(EM1))]
    M2obs = [map(lambda x:x+baselineM2[i], EM2[i]) for i in xrange(len(EM2))]
    M3obs = [map(lambda x:x+baselineM3[i], EM3[i]) for i in xrange(len(EM3))]
    if 'M4' in list(set(cpdReport['M'].tolist())):
        M4obs = [map(lambda x:x+baselineM4[i], EM4[i]) for i in xrange(len(EM4))]
    
    # Perform Offset Correction
    if len(offsetSlope) == 1:
        fitSlope = [[offsetSlope[0]]*len(sampleList)]*len(EM0)
        fitIntercept = [[offsetIntercept[0]]*len(sampleList)]*len(EM0)
    elif len(offsetSlope) == 4:
        fitSlope = map(list,zip(*[map(lambda x:offsetSlope[0]*np.log2(x)**3 + offsetSlope[1]*np.log2(x)**2 + offsetSlope[2]*np.log2(x) + offsetSlope[3], sumAbund[i]) for i in xrange(len(sampleList))]))
        fitIntercept = map(list,zip(*[map(lambda x:offsetIntercept[0]*np.log2(x)**3 + offsetIntercept[1]*np.log2(x)**2 + offsetIntercept[2]*np.log2(x) + offsetIntercept[3], sumAbund[i]) for i in xrange(len(sampleList))]))
    elif len(offsetSlope) == 0:
        fitSlope = [1]*len(EM0)
        fitIntercept = [0]*len(EM0)  
    
    # Re-Calculate Corrected Mx with Offset Correction
    M0cor = [map(lambda x,y,z:((x+baselineM0[i])-y)/z, EM0[i],fitIntercept[i],fitSlope[i]) for i in xrange(len(EM0))]
    M1cor = [map(lambda x,y,z:((x+baselineM1[i])-y)/z, EM1[i],fitIntercept[i],fitSlope[i]) for i in xrange(len(EM1))]
    M2cor = [map(lambda x,y,z:((x+baselineM2[i])-y)/z, EM2[i],fitIntercept[i],fitSlope[i]) for i in xrange(len(EM2))]
    M3cor = [map(lambda x,y,z:((x+baselineM3[i])-y)/z, EM3[i],fitIntercept[i],fitSlope[i]) for i in xrange(len(EM3))]
    if 'M4' in list(set(cpdReport['M'].tolist())):
        M4cor = [map(lambda x,y,z:((x+baselineM4[i])-y)/z, EM4[i],fitIntercept[i],fitSlope[i]) for i in xrange(len(EM4))]
    
    def squareFunc(a):
        return a**2
    
    # Re-Calculate EMx with Offset Correction
    EM0 = [map(lambda x:x-baselineM0[i], M0cor[i]) for i in xrange(len(M0cor))]
    EM1 = [map(lambda x:x-baselineM1[i], M1cor[i]) for i in xrange(len(M1cor))]
    EM2 = [map(lambda x:x-baselineM2[i], M2cor[i]) for i in xrange(len(M2cor))]
    EM3 = [map(lambda x:x-baselineM3[i], M3cor[i]) for i in xrange(len(M3cor))]
    if 'M4' in list(set(cpdReport['M'].tolist())):
        EM4 = [map(lambda x:x-baselineM4[i], M4cor[i]) for i in xrange(len(M4cor))]
    
    # Re-Calculate Theoretical EMx with Offset Correction
    theoEM0 = [[c/t if t!=0 else np.nan for c,t in zip(EM0[i],midaEM0[i])] for i in xrange(len(EM0))]
    theoEM1 = [[c/t if t!=0 else np.nan for c,t in zip(EM1[i],midaEM1[i])] for i in xrange(len(EM1))]
    theoEM2 = [[c/t if t!=0 else np.nan for c,t in zip(EM2[i],midaEM2[i])] for i in xrange(len(EM2))]
    theoEM3 = [[c/t if t!=0 else np.nan for c,t in zip(EM3[i],midaEM3[i])] for i in xrange(len(EM3))]
    if 'M4' in list(set(cpdReport['M'].tolist())):
        theoEM4 = [[c/t if t!=0 else np.nan for c,t in zip(EM4[i],midaEM4[i])] for i in xrange(len(EM4))]

    # Calculate RMS Error
    rmsError = []
    dummyMass = zip(*[list(zip(*pivotData2['Mass',sampleList[i]].values)[0]) for i in xrange(len(sampleList))])
    for i in xrange(len(EM0)):
        dummyData = zip(*[EM0[i],EM1[i],EM2[i],EM3[i],EM4[i]])         
        rmsError.append([np.sqrt(np.nansum(map(squareFunc,dummyData[j]))/5) if dummyMass[i][j]>2400 else np.sqrt(np.nansum(map(squareFunc,dummyData[j][0:4]))/4) for j in xrange(len(dummyData))])    
    
    # Calculate Retention Time
    newRt = map(list,zip(*[[pivotData2['Rt',sampleList[i]].iloc[j][3] for i in xrange(len(sampleList))] for j in xrange(len(pivotData2))]))
        
    ######################
    # Define Dictionary to output data
    if 'M4' in list(set(cpdReport['M'].tolist())):
        columnsToInsert = OrderedDict([
            ('RMS Error', map(list,zip(*rmsError))),
            ('RT', newRt),
            ('Sum Abund', sumAbund),
            ('sat', sat),
            ('Base Pk Abund', map(list,zip(*basePkAbund))),
            ('EM0', map(list,zip(*EM0))),
            ('%Theo EM0', map(list,zip(*theoEM0))),
            ('MIDA EM0', map(list,zip(*midaEM0))),
            ('EM1', map(list,zip(*EM1))),
            ('%Theo EM1', map(list,zip(*theoEM1))),
            ('MIDA EM1', map(list,zip(*midaEM1))),
            ('EM2', map(list,zip(*EM2))),
            ('%Theo EM2', map(list,zip(*theoEM2))),
            ('MIDA EM2', map(list,zip(*midaEM2))),
            ('EM3', map(list,zip(*EM3))),
            ('%Theo EM3', map(list,zip(*theoEM3))),
            ('MIDA EM3', map(list,zip(*midaEM3))),
            ('EM4', map(list,zip(*EM4))),
            ('%Theo EM4', map(list,zip(*theoEM4))),
            ('MIDA EM4', map(list,zip(*midaEM4))),
            ('M0_obs', map(list,zip(*M0obs))),
            ('M1_obs', map(list,zip(*M1obs))),
            ('M2_obs', map(list,zip(*M2obs))),
            ('M3_obs', map(list,zip(*M3obs))),
            ('M4_obs', map(list,zip(*M4obs))),
            ('M0_cor', map(list,zip(*M0cor))),
            ('M1_cor', map(list,zip(*M1cor))),
            ('M2_cor', map(list,zip(*M2cor))),
            ('M3_cor', map(list,zip(*M3cor))),
            ('M4_cor', map(list,zip(*M4cor)))])
    else:
        columnsToInsert = OrderedDict([
            ('RMS Error', map(list,zip(*rmsError))),
            ('RT', newRt),
            ('Sum Abund', sumAbund),
            ('sat', sat),
            ('Base Pk Abund', map(list,zip(*basePkAbund))),
            ('EM0', map(list,zip(*EM0))),
            ('%Theo EM0', map(list,zip(*theoEM0))),
            ('MIDA EM0', map(list,zip(*midaEM0))),
            ('EM1', map(list,zip(*EM1))),
            ('%Theo EM1', map(list,zip(*theoEM1))),
            ('MIDA EM1', map(list,zip(*midaEM1))),
            ('EM2', map(list,zip(*EM2))),
            ('%Theo EM2', map(list,zip(*theoEM2))),
            ('MIDA EM2', map(list,zip(*midaEM2))),
            ('EM3', map(list,zip(*EM3))),
            ('%Theo EM3', map(list,zip(*theoEM3))),
            ('MIDA EM3', map(list,zip(*midaEM3))),
            ('M0_obs', map(list,zip(*M0obs))),
            ('M1_obs', map(list,zip(*M1obs))),
            ('M2_obs', map(list,zip(*M2obs))),
            ('M3_obs', map(list,zip(*M3obs))),
            ('M0_cor', map(list,zip(*M0cor))),
            ('M1_cor', map(list,zip(*M1cor))),
            ('M2_cor', map(list,zip(*M2cor))),
            ('M3_cor', map(list,zip(*M3cor)))])
    
    for i in xrange(len(columnsToInsert)):
        for j in xrange(len(sampleList)):
            summaryTable[columnsToInsert.items()[i][0]+' - '+sampleList[j]] = columnsToInsert.items()[i][1][j]
    
    for i in xrange(len(summaryTable2.columns)):
        summaryTable.insert(len(summaryTable.columns),str(summaryTable2.columns[i]),summaryTable2[summaryTable2.columns[i]])
    
    # Add body wter data to dataframe
    for i in xrange(len(sampleList)):
        headers = ['BW'+' - '+sampleList[j] for j in xrange(len(sampleList))]
        summaryTable[headers[i]] =  map(list,zip(*[bwList] * len(pivotData2)))[i]  
    
    summaryTable = summaryTable.rename(columns = {'All Swissprot IDs':'All Acc #s','AAstart+seq':'Sequence'})
    
    summaryTablev2 = summaryTable[summaryTable['multiple IDs (1=yes)']==0]
    return summaryTable, map(list,zip(*basePkAbund)), summaryTablev2,sampleList

# Generate Data Filter   
def generateDataFilter(sTable, sampleList, basePkAbund,basePkAbundFilter,em0UpperLimit,em0LowerLimit,isotopomerSDFilter,minMIDAforEMx,peptideSDFilter,RMSEFilter,useAllIsotopomers,saturation):
    dataFilter = sTable.ix[0:,3:13]
    dataFilter.columns = [str(i) for i in dataFilter.columns]
    
    df2 = sTable.copy()
    # Extract Isotopomer data
    
    isotopomerData = [[i for i in sTable[[j for j in sTable.columns if j.find(str(sampleList[k]))!=-1 and j.find("%Theo")!=-1]].values] for k in xrange(len(sampleList))]
    #isotopomerData = [[i for i in sTable[[j for j in sTable.columns if j.find(str(sampleList[k]))!=-1 and j.find("EM")!=-1 and j.find("EM")!=-1 and j.find('MIDA')==-1 and j.find('cubic')==-1 and j.find('Theo')==-1]].values] for k in xrange(len(sampleList))]
    
    ### Optional code to perform calculations for control samples
    #ControlRMSE = [i for i in sTable[[j for j in sTable.columns if j.find("Control")!=-1 and j.find("RMS Error")!=-1]].values]
    
    #minControlRMSE = [[np.nanmin([i for i in ControlRMSE[j] if i>0]) if len([i for i in ControlRMSE[j] if i>0])>0 else np.nan] for j in xrange(len(ControlRMSE))]

    #if len([i for i in sampleList if i.upper().find("CONTROL")!=-1])>0:
    ##    if len([i for i in sampleList if i.upper().find("CONTROL")!=-1])>0:
    ##        minControlRMSE = map(np.nanmin,[i for i in sTable[[j for j in sTable.columns if j.upper().find("CONTROL")!=-1 and j.find("RMS Error")!=-1]].values])
    ##    else:
    minControlRMSE = [np.nan]*len(isotopomerData[0])
    
    isoMIDA = [[i for i in sTable[[j for j in sTable.columns if j.find(sampleList[k])!=-1 and j.find("MIDA")!=-1]].values] for k in xrange(len(sampleList))]
    
    # Filter isotopomer data
    for i in xrange(len(isotopomerData)):
        for j in xrange(len(isotopomerData[i])):
            if basePkAbund[i][j]<basePkAbundFilter or saturation[i][j]==1 or np.greater(isotopomerData[i][j][0],em0UpperLimit) or np.less_equal(isotopomerData[i][j][0],em0LowerLimit) or minControlRMSE[j]>=RMSEFilter:
                isotopomerData[i][j][0:] = [np.nan]*len(isotopomerData[i][j])
            elif abs(isoMIDA[i][j][1])<abs(minMIDAforEMx[i]) or saturation[i][j]==1:
                isotopomerData[i][j][1] = np.nan
            elif abs(isoMIDA[i][j][2])<abs(minMIDAforEMx[i]) or saturation[i][j]==1:
                isotopomerData[i][j][2] = np.nan
            elif abs(isoMIDA[i][j][3])<abs(minMIDAforEMx[i]) or saturation[i][j]==1:
                isotopomerData[i][j][3] = np.nan
    
    # Calculate isotopomer distribution    
    isotopomerSD = [[np.std([a for a,a in enumerate(isotopomerData[j][i]) if np.isnan(a)==False],ddof=1)  for i in xrange(len(isotopomerData[j]))] for j in xrange(len(isotopomerData))]
    
    if useAllIsotopomers ==False:
        filteredAvgF = [[isotopomerData[j][i][0] if isotopomerSD[j][i]<isotopomerSDFilter else np.nan for i in xrange(len(isotopomerData[j]))] for j in xrange(len(isotopomerData))]
        filteredAvgFraw = [[isotopomerData[j][i][0] for i in xrange(len(isotopomerData[j]))] for j in xrange(len(isotopomerData))]
        
    if useAllIsotopomers ==True:
        filteredAvgF = [[np.mean([a for a,a in enumerate(isotopomerData[j][i]) if np.isnan(a)==False]) if isotopomerSD[j][i]<isotopomerSDFilter else np.nan for i in xrange(len(isotopomerData[j]))] for j in xrange(len(isotopomerData))]
        filteredAvgFraw = [[np.mean([a for a,a in enumerate(isotopomerData[j][i]) if np.isnan(a)==False]) for i in xrange(len(isotopomerData[j]))] for j in xrange(len(isotopomerData))]
        
    for i in xrange(len(sampleList)):
        headers = ['Filtered Avg f'+' - '+sampleList[j] for j in xrange(len(sampleList))]
        df2[headers[i]] = filteredAvgF[i]

    for i in xrange(len(sampleList)):
        headers1 = ['Avg f'+' - '+sampleList[j] for j in xrange(len(sampleList))]
        df2[headers1[i]] = filteredAvgFraw[i]
    
    for i in xrange(len(sampleList)):
        headers1b = ['Isotopomer SD'+' - '+sampleList[j] for j in xrange(len(sampleList))]
        df2[headers1b[i]] = isotopomerSD[i]
        
    def getSDF(x):
        return [i for i in x if np.isnan(i)==False][0] if len([i for i in x if np.isnan(i)==False])==1 else np.std([i for i in x if np.isnan(i)==False],ddof=1)

    # Calculate Protein Average f and standard deviation
    accessionNumbers = np.sort(list(set(df2['Accession#'])))
    proteinAverageF = pd.DataFrame(data=map(list,zip(*[df2[headers[i]].groupby(df2['Accession#'],sort=True).mean().values for i in xrange(len(headers))])),index=accessionNumbers,columns=['Protein Avg f'+' - '+sampleList[j] for j in xrange(len(sampleList))])
    proteinStdDevF = pd.DataFrame(data=map(list,zip(*[map(getSDF,df2[headers[i]].groupby(df2['Accession#'],sort=True).values) for i in xrange(len(headers))])),index=accessionNumbers,columns=['Peptide SD f'+' - '+sampleList[j] for j in xrange(len(sampleList))])
    testDataFrame = pd.merge(df2,proteinAverageF,left_on='Accession#',right_index=True,how='left',sort=False)
    testDataFrame2 = pd.merge(testDataFrame,proteinStdDevF,left_on='Accession#',right_index=True,how='left',sort=False)

    headers2 = ['Protein Avg f'+' - '+sampleList[j] for j in xrange(len(sampleList))]
    headers3 = ['Peptide SD f'+' - '+sampleList[j] for j in xrange(len(sampleList))]

    # Calculate Filtered Average f
    filteredF2 = [[testDataFrame2[headers[i]][j] if np.less_equal(testDataFrame2[headers[i]][j],testDataFrame2[headers2[i]][j]+(peptideSDFilter*testDataFrame2[headers3[i]][j])) and np.greater_equal(testDataFrame2[headers[i]][j],testDataFrame2[headers2[i]][j]-(peptideSDFilter*testDataFrame2[headers3[i]][j])) else np.nan for i in xrange(len(sampleList))] for j in xrange(len(testDataFrame2))]

    dummy1 = map(list,zip(*filteredF2))
    for i in xrange(len(sampleList)):
        headers4 = ['Filtered Avg f2'+' - '+sampleList[j] for j in xrange(len(sampleList))]
        testDataFrame2[headers4[i]] = dummy1[i]

    groupedByProtein = pd.DataFrame(testDataFrame2.groupby(testDataFrame2['Accession#'],axis=0,sort=False).first())
    groupedByProtein = groupedByProtein.drop(['n','Charge'],axis=1).reset_index(drop=True)

    averageF = [testDataFrame2[headers4[i]].groupby(testDataFrame2['Accession#'],sort=False).median().round(3).values for i in xrange(len(headers))]
    stddevF = [testDataFrame2[headers4[i]].groupby(testDataFrame2['Accession#'],sort=False).std().round(3).values for i in xrange(len(headers))]
    countF = [testDataFrame2[headers4[i]].groupby(testDataFrame2['Accession#'],sort=False).count().values for i in xrange(len(headers))]
    
    proteinColumns = OrderedDict([
        ('Protein f',averageF),
        ('Protein SD',stddevF),
        ('Count',countF)])
        
    for i in xrange(len(proteinColumns)):
        for j in xrange(len(sampleList)):
            groupedByProtein[proteinColumns.items()[i][0]+' - '+sampleList[j]] = proteinColumns.items()[i][1][j]

    del groupedByProtein['Sequence']
    del groupedByProtein['Modifications']
    del groupedByProtein['Mass']
    del groupedByProtein['Fraction']    
    
    groupedByProtein.insert(4,'Mean Peptide Count',[np.round(np.mean([x for x in map(list,zip(*countF))[i] if x>0]),1) for i in xrange(len(groupedByProtein))])
    ##    if len([i for i in sampleList if i.upper().find("CONTROL")!=-1])>0:
    ##        testDataFrame2.insert(11,'Min Control RMSE',minControlRMSE)
    ##        peptideColumns = ['Species','Code','Protein','Accession#','Sequence','Modifications','Mass','n','Charge','Fraction','Target RT','Min Control RMSE']+headers4
    ##    else:
    
    #peptideColumns = ['Species','Code','Protein','Accession#','Sequence','Modifications','Mass','n','Charge','Fraction','Target RT']+headers4
    peptideColumns = ['Species','Code','Protein','Accession#','Sequence','Modifications','Mass','n','Charge','Fraction','Target RT']+headers4+headers1b

    headers5 = ['Protein f'+' - '+sampleList[j] for j in xrange(len(sampleList))]
    headers6 = ['Protein SD'+' - '+sampleList[j] for j in xrange(len(sampleList))]
    headers7 = ['Count'+' - '+sampleList[j] for j in xrange(len(sampleList))]
    proteinColumns = ['Species','Code','Protein','Accession#','Target RT','Mean Peptide Count']+headers5+headers6+headers7+[str(i) for i in sTable.columns if i.startswith('BW - ')]
    
    peptideTab = testDataFrame2.ix[0:,peptideColumns].sort(columns=['Protein','Sequence','Charge'])
    peptideTab.Mass = peptideTab.Mass.round(2)
    for i in xrange(len(headers4)):
        peptideTab[headers4[i]] = peptideTab[headers4[i]].round(3)
        
    for i in xrange(len(headers1b)):
        peptideTab[headers1b[i]] = peptideTab[headers1b[i]].round(3)    
    
    peptideTab.columns = ['Species','Code','Protein','Accession#','Sequence','Modifications','Mass','n','Charge','Fraction','Target RT']+headers+headers1b
    proteinTab = groupedByProtein.ix[0:,proteinColumns].sort(columns=['Protein'])
    
    return peptideTab, proteinTab

# Generate and export 'Analysis Summary' pdf file
import fpdf
def exportParameterPDF(PDFoutputFileName,acquiredDate,instrument,today,sampleInput, 
    bodyWaterData, rtDiffFilter, totalAbundFilter, mDiffCriteria, silacMasses, 
    rmsErrorFilter, dbScoreFilter, basePkAbundFilter, em0UpperLimit, em0LowerLimit, 
    peptideSDFilter, isotopomerSDFilter,midaFiles,crFiles,projectLeader,
    processedBy,notebookCode,projectCode,scriptVersion,minMIDAEMx,upSILAC, lowSILAC, 
    silacStdDev,useAllIsotopomers, combinePeptides, useFractions, minMIDAarray,
    satLimit,offsetSlope,offsetIntercept):
    
    parameterPDF = fpdf.FPDF(format = 'letter')
    parameterPDF.add_page()
    parameterPDF.set_font('Arial','BIU', size=14)
    parameterPDF.cell(200,10, txt="Analysis Summary", align="C")
    parameterPDF.cell(200,10,'',0,1,'L')

    parameterPDF.set_font('Arial','BI',size=12)
    parameterPDF.cell(200,5,'SUBMISSION DETAILS','B','L')
    parameterPDF.cell(200,5,'',0,1,'L')
    parameterPDF.set_font('Arial',size=12)
    
    analysisNames = [
        'Analyzed By',
        'Analysis Date',
        'Script Version',
        'Acquired Date',
        'Project Code',
        'Instrument',
        'Project Leader',
        'Processed By',
        'Notebook Code']
    
    analysisValues = [
        os.environ.get("USERNAME"),
        today.strftime("%m/%d/%Y"),
        scriptVersion,
        acquiredDate,
        projectCode,
        instrument,
        projectLeader,
        processedBy,
        notebookCode]
    
    analysisData = map(list,zip(analysisNames,analysisValues))

    for i in xrange(len(analysisData)):
        for j in xrange(len(analysisData[i])):
            parameterPDF.cell(75,7.5,str(analysisData[i][j]),1,0)
        parameterPDF.ln()
    parameterPDF.cell(200,7.5,'',0,1,'L')

    parameterPDF.set_font('Arial','BI',size=12)
    parameterPDF.cell(200,5,'FILES','B','L')
    parameterPDF.cell(50,5,'',0,1,'L')

    parameterPDF.set_font('Arial','I',size=12)
    parameterPDF.cell(200,7.5,'MIDA Database(s):',1,0)
    parameterPDF.ln()
    parameterPDF.set_font('Arial',size=6)
    for i in xrange(len(midaFiles)):
        if i == xrange(len(midaFiles))[-1]:
            parameterPDF.cell(200,5,midaFiles[i],'LRB',0)
        else:
            parameterPDF.cell(200,5,midaFiles[i],'LR',0)
        parameterPDF.ln()
    parameterPDF.cell(200,5,'',0,1,'L')

    parameterPDF.set_font('Arial','I',size=12)
    parameterPDF.cell(200,7.5,'Compound Reports: ',1,0)
    parameterPDF.ln()
    parameterPDF.set_font('Arial',size=6)
    for i in xrange(len(crFiles)):
        if i == xrange(len(crFiles))[-1]:
            parameterPDF.cell(200,5,crFiles[i],'LRB',0)
        else:
            parameterPDF.cell(200,5,crFiles[i],'LR',0)
        parameterPDF.ln()
    parameterPDF.cell(200,10,'',0,1,'L')

    parameterPDF.set_font('Arial','BI',size=12)
    parameterPDF.cell(200,5,'ANALYSIS DETAILS','B','L')
    parameterPDF.cell(200,5,'',0,1,'L')
    
    headers = ['Sample', 'Body Water', 'Min. MIDA for EMx']
    for i in xrange(len(headers)):
        parameterPDF.cell(75,7.5,headers[i],1,0)
    parameterPDF.ln()
    parameterPDF.set_font('Arial',size=12)
    bodyWaterDataPercent = [str(round(bodyWaterData[i]*100,3)) +'%' for i in xrange(len(bodyWaterData))]
    midaList = [str(round(minMIDAarray[i], 3)) for i in xrange(len(minMIDAarray))]
    chartData = map(list,set(zip(sampleInput,bodyWaterDataPercent,midaList)))
    chartData.sort()
    for i in xrange(len(chartData)):
        for j in xrange(len(chartData[i])):
            parameterPDF.cell(75,7.5,str(chartData[i][j]),1,0)
        parameterPDF.ln()
    parameterPDF.cell(200,10,'',0,1,'L')
    
    parameterPDF.set_font('Arial','BI',size=12)
    parameterPDF.cell(200,5,'PARAMETERS','B','L')
    parameterPDF.cell(50,5,'',0,1,'L')
    headers2 = ['Parameter', 'Value']
    for i in xrange(len(headers2)):
        parameterPDF.cell(75,7.5,headers2[i],1,0)
    parameterPDF.ln()
    parameterPDF.set_font('Arial',size=12)

    parameterNames = [
        'Rt Difference Filter',
        'Total Abundance Filter',
        'RMS Error Filter',
        'DB Score Filter',
        'Base Peak Abundance Filter',
        '% Theo EM0 Upper Limit',
        '% Theo EM0 Lower Limit',
        'Peptide Std Dev Filter',
        'Isotopomer Std Dev Filter',
        'Use All Isotopomers for f',
        'Combine peptides',
        'Min. MIDA for EMx',
        'SILAC Masses',
        'Sample:SILAC Ratio Upper Limit',
        'Sample:SILAC Ratio Upper Limit',
        'SILAC Std Dev Filter',
        'Fractionated Samples',
        'Saturation Limit',
        'Offset Correction Slope',
        'Offset Correction Intercept']
        
    parameterValues = [
        str(rtDiffFilter),
        str(int(totalAbundFilter)),
        str(rmsErrorFilter),
        str(dbScoreFilter),
        str(int(basePkAbundFilter)),
        str(em0UpperLimit),
        str(em0LowerLimit),
        str(peptideSDFilter),
        str(isotopomerSDFilter),
        "Yes" if str(useAllIsotopomers)=='True' else "No",
        "Yes" if str(combinePeptides)=='True' else "No",
        str(minMIDAEMx),
        str(silacMasses),
        str(upSILAC), 
        str(lowSILAC), 
        str(silacStdDev),
        "Yes" if str(useFractions)=='True' else "No",
        str(satLimit),
        str(offsetSlope),
        str(offsetIntercept)]
        
    parameterData = map(list,zip(parameterNames,parameterValues))
    for i in xrange(len(parameterData)):
        for j in xrange(len(parameterData[i])):
            parameterPDF.cell(75,7.5,parameterData[i][j],1,0)
        parameterPDF.ln()
    
    parameterPDF.output(PDFoutputFileName)

# Check Sample Dictionary to ensure samples match Compound Reports
def checkDict(currentDict):
    errors = []
    for i in currentDict:
        if i[0:5]!=currentDict[i][-5:]:
            errors.append(0)
    if len(errors)>0:
        print '*** SAMPLE NAMES FROM SUBMISSION FORM DO NOT MATCH COMPOUND REPORTS! ***'
        print '*** CHECK OUTPUT FILES ***'

# Format final Excel files
from openpyxl.cell import get_column_letter
from openpyxl.style import Color, Fill

def formatExcel(fileName,numSheets,nPeptides,nProteins,nSamples,workSheetA,workSheetB):
    
    Color.LightPurple = 'E4DFEC'
    Color.DarkBlue = '00008B'
    Color.DarkPurple=	'CCC0DA'
    
    wb = load_workbook(fileName)
    ws = wb.worksheets[workSheetA]
    ws._freeze_panes = 'A2'
    ws.cell('A1').style.alignment.wrap_text = True
    ws.row_dimensions[1].height = 60.0
    peptideRange = 'L2:'+get_column_letter(nSamples+11)+str(nPeptides+1)
    for row in ws.range(peptideRange):
        for cell in row:
            cell.style.number_format.format_code = '0.000'
    
    if numSheets>1:
        ws2 = wb.worksheets[workSheetB]
        ws2._freeze_panes = 'A2'
        ws2.cell('A1').style.alignment.wrap_text = True
        ws2.row_dimensions[1].height = 60.0
        proteinRange1 = 'A2:'+get_column_letter((nSamples)+6)+str(nProteins+1)
        for column in ws2.range(proteinRange1):
            for cell in column:
                if cell.column=='F':
                    cell.style.number_format.format_code = '0.0'
                elif cell.column=='E':
                    cell.style.number_format.format_code = '0.00'
                else:
                    cell.style.number_format.format_code = '0.000'
                if cell.column=='F':
                    cell.style.font.color.index = Color.RED
                else:
                    cell.style.font.color.index = Color.DarkBlue
                cell.style.fill.fill_type = Fill.FILL_SOLID
                cell.style.fill.start_color.index = Color.LightPurple
        proteinRange2 = get_column_letter((nSamples)+7)+'2:'+get_column_letter((2*nSamples)+6)+str(nProteins+1)
        for row in ws2.range(proteinRange2):
            for cell in row:
                cell.style.number_format.format_code = '0.000'
                cell.style.font.color.index = Color.DarkBlue
                cell.style.fill.fill_type = Fill.FILL_SOLID
                cell.style.fill.start_color.index = Color.DarkPurple
        proteinRange3 = get_column_letter(ws2.get_highest_column())+'2:'+get_column_letter(ws2.get_highest_column())+str(ws2.get_highest_row())
        for row in ws2.range(proteinRange3):
            for cell in row:
                cell.hyperlink = cell.value
                cell.style.font.color.index = Color.BLUE
    
    wb.save(fileName)

# Functions to match and sort compound report to sample names
def getFullSamples(fName,sName):
    c = ['']*len(fName)
    for i in range(len(fName)):
        for j in range(len(sName)):
            if fName[i].find(sName[j][0:len(sName[j])-sName[j][::-1].find('-')-1])!=-1 and fName[i].find(sName[j][-5:])!=-1 and sName[j].find(fName[i][fName[i].find('_D')+2:fName[i].find('_MS')])!=-1:
                c[i]= sName[j]
    return c
    
def getFullSamples2(fName,sName):
    dummy = []
    dummy2 = []
    for i in xrange(len(sName)):
        for j in xrange(len(fName)):
            if sName[i][-5:]==fName[j][0:5] and sName[i][0:(len(sName[i])-1-sName[i][::-1].find('-'))]==fName[j][::-1][fName[j][::-1].find('SM_')+3:fName[j][::-1].find('D_')][::-1]:
                dummy.append(sName[i])
                dummy2.append(fName[j])
    return dummy, dummy2

################################################################################

# Main loop
def main(fileLoc):
    
    today = datetime.today()
    scriptVersion = "Script_v5.0"
    
    print 'IMPORTING DATA...'
    # Import data and parameters
    parametersData, midaDatabases,compoundReports = getFiles(fileLoc)
    sampleInput, bodyWaterData, rtDiffFilter, totalAbundFilter, mDiffCriteria, silacMasses, rmsErrorFilter, dbScoreFilter, basePkAbundFilter, em0UpperLimit, em0LowerLimit, peptideSDFilter, isotopomerSDFilter, useAllIsotopomers, combinePeptides, instrument, projectLeader, processedBy, submitDate, projectCode, notebookCode, tissueFluid, prep, minMIDAEMx, upperSILAC, lowerSILAC, silacSD, fractionated, minMIDAstring, saturationLevel, correctionSlope, correctionIntercept = loadParameters(parametersData)
    
    # Sort Compound Report files
    compoundReports.sort(key=sortFiles2)
    
    if os.name == 'nt':
        crFileSet = [i.split('\\')[-1][0:i.split('\\')[-1].find('Compound')-1]+'.d' for i in compoundReports]
    elif os.name == 'posix':
        crFileSet = [i.split('/')[-1][0:i.split('/')[-1].find('Compound')-1]+'.d' for i in compoundReports]
    newSamples, newCRFiles = getFullSamples2(crFileSet,sampleInput)
    sampleDict = OrderedDict((x,y) for (x,y) in zip(newCRFiles,newSamples))
    
    print 'CONVERTING COMPOUND REPORTS...'
    # Convert Compound Report Files to sorted and filtered dataframe for subsequent calculations
    crOutput, crSummaryOutput, crSorted, acquiredDate = convertCompoundReport(compoundReports,sampleDict,mDiffCriteria,totalAbundFilter,fractionated)
    
    def checkDate(inDate):
        dummyDate = inDate.replace('\\','-').replace('/','-')
        if len(dummyDate[0:dummyDate.find('-')])==1:
            dummyDate = '0'+dummyDate
        if len(dummyDate[dummyDate.find('-')+1:][0:dummyDate[dummyDate.find('-')+1:].find('-')])==1:
            dummyDate = dummyDate[0:dummyDate.find('-')]+'-0'+dummyDate[-6:]
        return dummyDate                
    
    dateInstrument = checkDate(acquiredDate)+instrument
    
    # Define output filenames
    SToutputFileName = [i for i in crOutput.Code if i!=0][0]+'_'+list(set(crOutput.Fraction))[0]+'_'+dateInstrument+'_'+'SummaryTable_'+scriptVersion+'_done'+today.strftime("%m%d%Y")+'.xlsx'
    DFoutputFileName = [i for i in crOutput.Code if i!=0][0]+'_'+list(set(crOutput.Fraction))[0]+'_'+dateInstrument+'_'+'DataFilter_'+scriptVersion+'_done'+today.strftime("%m%d%Y")+'.xlsx'
    PDFoutputFileName = [i for i in crOutput.Code if i!=0][0]+'_'+list(set(crOutput.Fraction))[0]+'_'+dateInstrument+'_'+'AnalysisSummary'+'_done'+today.strftime("%m%d%Y")+'.pdf'
    
    # END COMPOUND REPORT CONVERTER
    ######################################################################################################################
    # BEGIN SUMMARY FILE 
    print 'GENERATING SUMMARY TABLE...'
    # Calculate and export Summary Table
    stSummary, basePeakAbundance, stSummaryV2,newSampleList = generateSummaryFile(sampleInput,midaDatabases,crSorted,bodyWaterData,silacMasses,saturationLevel,correctionSlope, correctionIntercept)
    stSummary2 = stSummary.groupby(by=['Sample Score','Abund Score','Sample*Abund Score','Species','Code','Protein','Accession#','Sequence','Mass','n','Charge','Fraction'])
    stSummary3 = stSummary2.agg(np.max).reset_index()
    stSummary4 = stSummary3.sort(columns=['Protein','Sequence','Charge'])
    stSummary4.to_excel(SToutputFileName, sheet_name = 'SUMMARY TABLE',index=False)
    
    # END SUMMARY FILE
    #############################################################################################
    bPA = map(list,zip(*stSummary3[[j for j in stSummary4.columns if j.find("Base Pk Abund")!=-1]].values))
    sat = map(list,zip(*stSummary3[[j for j in stSummary4.columns if j.find("sat")!=-1]].values))
    # BEGIN DATA FILTER
    print 'GENERATING DATA FILTER...'
    dfPeptides, dfProteins = generateDataFilter(stSummary3, newSampleList, bPA,basePkAbundFilter,em0UpperLimit,em0LowerLimit,isotopomerSDFilter,minMIDAEMx,peptideSDFilter,rmsErrorFilter,useAllIsotopomers,sat)
    
    parameterNames = [
        'Rt Difference Filter',
        'Total Abundance Filter',
        'M Difference Criteria',
        'SILAC Masses',
        'RMS Error Filter',
        'DB Score Filter',
        'Base Peak Abundance Filter',
        '%Theo EM0 Upper Limit',
        '%Theo EM0 Lower Limit',
        'Peptides Std Dev Filter',
        'Isotopomer Std Dev Filter',
        'Use All Isotopomers for f',
        'Combine peptides',
        'Minimum MIDA',
        'Upper limit Sample:SILAC ratio',
        'Lower limit Sample:SILAC ratio',
        'SILAC Std Dev Filter',
        'Saturation Limit',
        'Offset Correction Slope',
        'Offset Correction Intercept']

    parameterValues = [
        rtDiffFilter,
        totalAbundFilter,
        ', '.join([str(i) for i in mDiffCriteria]),
        silacMasses,
        rmsErrorFilter,
        dbScoreFilter,
        basePkAbundFilter,
        em0UpperLimit,
        em0LowerLimit,
        peptideSDFilter,
        isotopomerSDFilter,
        useAllIsotopomers,
        combinePeptides,
        minMIDAstring,
        upperSILAC,
        lowerSILAC,
        silacSD,
        saturationLevel,
        ', '.join([str(i) for i in correctionSlope]), 
        ', '.join([str(i) for i in correctionIntercept])]

    projectDetailNames = [
        'Project Leader',
        'Processed By',
        'Project Code',
        'Notebook Code',
        'Tissue/Fluid',
        'Prep']
        
    projectDetailValues = [
        projectLeader,
        processedBy,
        projectCode,
        notebookCode,
        tissueFluid,
        prep
        ]
                
    parameterData = pd.DataFrame({'Parameter':parameterNames,'Value':parameterValues})
    projectData = pd.DataFrame({'Parameter':projectDetailNames,'Value':projectDetailValues})
    dfProteins = dfProteins.reset_index()
    dfProteins['Uniprot Link'] = [str('http://www.uniprot.org/uniprot/'+dfProteins['Accession#'][i]) for i in xrange(len(dfProteins))]
    del dfProteins['index']
    
    writer2 = pd.ExcelWriter(DFoutputFileName) 
    dfPeptides.to_excel(writer2,'PEPTIDE OUTPUT',index=False) 
    dfProteins.to_excel(writer2,'PROTEIN OUTPUT',index=False)
    parameterData.to_excel(writer2,'PARAMETERS',index=False) 
    projectData.to_excel(writer2,'PROJECT DETAILS',index=False) 
    writer2.save()  
    #END DATA FILTER
    ################################################################################
    print 'GENERATING ANALYSIS SUMMARY...'
    exportParameterPDF(PDFoutputFileName,acquiredDate,instrument,today,sampleInput, bodyWaterData, 
    rtDiffFilter, totalAbundFilter, mDiffCriteria, silacMasses, rmsErrorFilter, 
    dbScoreFilter, basePkAbundFilter, em0UpperLimit, em0LowerLimit, 
    peptideSDFilter, isotopomerSDFilter,midaDatabases, compoundReports,
    projectLeader, processedBy,notebookCode,projectCode,scriptVersion,minMIDAstring,upperSILAC, 
    lowerSILAC, silacSD,useAllIsotopomers, combinePeptides,fractionated,minMIDAEMx,saturationLevel,
    correctionSlope, correctionIntercept)

    print 'FORMATTING OUTPUT...'
    formatExcel(DFoutputFileName,2,len(dfPeptides),len(dfProteins),len(list(set(sampleInput))),0,1)#fileName,numSheets,nPeptides,nProteins,nSamples
    
    textFile = list(set(crOutput.Code))[0]+'_'+list(set(crOutput.Fraction))[0]+'_'+dateInstrument+'_'+'Python'+scriptVersion+'_done'+today.strftime("%m%d%Y")+'.txt'
    
    return textFile, projectLeader, projectCode, processedBy, notebookCode, today.strftime("%Y"), midaDatabases[0], parametersData, DFoutputFileName, SToutputFileName, PDFoutputFileName

# Create and save email message to project leader and research associate
def CreateEmailMessage(ProjectLeader, ProcessedBy, Subject, FolderLocation1, FolderLocation2, Comments):
    
    Outlook = win32.Dispatch('outlook.application')
    Email = Outlook.CreateItem(0)
    ProjectLeaderEmail = ProjectLeader[0]+ProjectLeader.split(' ')[-1]+'@kinemed.com'
    ProcessedByEmail = ProcessedBy[0]+ProcessedBy.split(' ')[-1]+'@kinemed.com'
    
    Email.To = ';'.join([ProjectLeaderEmail,ProcessedByEmail])
    Email.CC = 'TAngel@kinemed.com'
    Email.Subject = Subject
    Email.GetInspector ### ADDS SIGNATURE TO EMAIL 
    index = Email.HtmlBody.find('>',Email.HtmlBody.find('<body>'))
    
    Link1 = '<a href='+urllib.quote(FolderLocation1, '\\')+'>'+FolderLocation1+'</a>'
    Link2 = '<a href='+urllib.quote(FolderLocation2, '\\')+'>'+FolderLocation2+'</a>'

    Notes = 'NOTE: ' + Comments
    if Comments == '':
        Email.HtmlBody = Email.HtmlBody[index+1:] + Link1 + "<br><br>" + Link2 + Email.HtmlBody[:index+1]
    else:
        Email.HtmlBody = Email.HtmlBody[index+1:] + Link1 + "<br><br>" + Link2 + "<br><br>" + Notes + Email.HtmlBody[:index+1]
        
    #Email.Display(True)
    Email.Save()
    Outlook.Quit()
    #Email.Send() #Optional code to automatically send message
    return
    
# Execute Script and write results to J drive folder
if __name__ == "__main__":
    src = os.path.dirname(os.path.abspath(__file__))
    np.warnings.simplefilter("ignore",FutureWarning)
    np.warnings.simplefilter("ignore",RuntimeWarning)
    #textOut = main(src)
    #try:
    start = time.time()
    # Execute script and return all output files and information
    Result, ProjectLeader, ProjectCode, ProcessedBy, NotebookCode, AnalysisYear, MidaDBFilename, ParametersFilename, DataFilterFilename, SummaryTableFilename, AnalysisSummaryFilename = main(src)
    
    # If successful, update data processing log    
    if os.name == 'nt':
        if os.path.exists('D:\\ResultsLog\\'): # PATH FOR DATA PROCESSING LOG
            with open('D:\\ResultsLog\\GOOD--'+Result, "a") as myfile:
                myfile.write('Script Successful'+'\n')
                myfile.write(src+'\n')
        
    print 'COPYING OUTPUT TO NETWORK...'
    # Copy folder/files to both ExtractedResult and LCMS Results folders on the network
    CurrentFolder = src.split('\\')[-1]
    NetworkDirectory = 'D:\\AnalyticalCore\\'
    CopyLocation1 = os.path.join(NetworkDirectory,'ExtractedResults',ProjectLeader[0]+ProjectLeader.split(' ')[-1],ProjectCode,AnalysisYear,CurrentFolder)
    CopyLocation2 = os.path.join(NetworkDirectory,'LCMS Results',ProjectLeader,ProjectCode,AnalysisYear,NotebookCode)

    # Create new folder on network drives
    # If folder/files already exist, append folder name with Version number
    if os.path.exists(CopyLocation1) == False:
        shutil.copytree(src,CopyLocation1)
    else:
        FolderList = [i[0] for i in os.walk(os.path.join(NetworkDirectory,'ExtractedResults',ProjectLeader[0]+ProjectLeader.split(' ')[-1],ProjectCode,AnalysisYear)) if i[0].find(CurrentFolder)!=-1]
        FolderList.sort()
        if FolderList[-1].find('Version')==-1:
            CurrentFolder = CurrentFolder+'_Version1'
        else:
            CurrentFolder = CurrentFolder+'_Version'+str(int(FolderList[-1].split('_')[-1].split('Version')[-1])+1)
        CopyLocation1 = os.path.join(NetworkDirectory,'ExtractedResults',ProjectLeader[0]+ProjectLeader.split(' ')[-1],ProjectCode,AnalysisYear,CurrentFolder)
        shutil.copytree(src,CopyLocation1)
        
    TextNotes = [n for n in [i for j in [k[2] for k in os.walk(src)] for i in j] if n.find('NOTE')!=-1 and n.endswith('.txt')]

    MidaDBFilename = MidaDBFilename.split('\\')[-1]
    ParametersFilename = ParametersFilename.split('\\')[-1]
    
    if os.path.exists(CopyLocation2) == False:
        os.makedirs(CopyLocation2)
    else:
        FolderList = [i[0] for i in os.walk(os.path.join(NetworkDirectory,'LCMS Results',ProjectLeader,ProjectCode,AnalysisYear)) if i[0].find(NotebookCode)!=-1]
        FolderList.sort()
        if FolderList[-1].find('Version')==-1:
            CurrentFolder = NotebookCode+'_Version1'
        else:
            CurrentFolder = NotebookCode+'_Version'+str(int(FolderList[-1].split('_')[-1].split('Version')[-1])+1)
        CopyLocation2 = os.path.join(NetworkDirectory,'LCMS Results',ProjectLeader,ProjectCode,AnalysisYear,CurrentFolder)
        os.makedirs(CopyLocation2)# ,mode=0777)

    shutil.copy(MidaDBFilename,CopyLocation2)
    shutil.copy(DataFilterFilename,CopyLocation2)
    shutil.copy(SummaryTableFilename,CopyLocation2)
    shutil.copy(AnalysisSummaryFilename,CopyLocation2)
    for i in xrange(len(TextNotes)):
        shutil.copy(TextNotes[i],CopyLocation2)
    
    #print 'CREATING COPIES FOR THE ATLAS...'
    # Create copies of all output files in deisgnated folder for the Atlas
##        AtlasDirectory = '\\\\kinemed-data\\Scientists\\Documents\\Marc Colangelo\\ADD_TO_ATLAS\\'
##        shutil.copy(MidaDBFilename,os.path.join(AtlasDirectory,'MidaDatabases'))
##        shutil.copy(DataFilterFilename,os.path.join(AtlasDirectory,'DataFilters'))
##        shutil.copy(SummaryTableFilename,os.path.join(AtlasDirectory,'SummaryTables'))
##        shutil.copy(ParametersFilename,os.path.join(AtlasDirectory,'ParameterTemplates'))
    
    #print 'CREATING EMAIL MESSAGE...'
    # Generate email message to project leader, sample processor
##        Comments = ''
##        CreateEmailMessage(ProjectLeader, ProcessedBy, CopyLocation1.split('\\')[-1], CopyLocation1, CopyLocation2, Comments)
    
    end = time.time()
    print 'COMPUTATION TIME: ',np.round(end-start,2)

    raw_input('SCRIPT COMPLETE!...PRESS ENTER TO EXIT')   
    # except Exception, err:
    #     # If script fails, updated log with folder name and error message
    #     badFile = src.split('\\')[-1]
    #     if os.name == 'nt':
    #         if os.path.exists('D:\\ResultsLog\\'):
    #             with open('D:\\ResultsLog\\BAD--'+badFile+'.txt', "a") as myfile:
    #                 myfile.write(str(err)+'\n')
    #                 myfile.write(src+'\n')
                
    #     raw_input('*** CHECK SAMPLE NAMES AND COMPOUND REPORT NAMES ***... HIT ENTER TO EXIT')


